import openpyxl
from openpyxl.styles import Font
import time

class ExcelParsing():
    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color=None)
        self.colDict = {'red':'FFFF3030',
                        'blue': '0000FF'}
        self.now = time.strftime('%Y-%m-%d %H:%M:%S')

    def getWorkBook(self, filePathAndName):
        self.workbook = openpyxl.load_workbook(filePathAndName)
        self.excelFile = filePathAndName
        return self.workbook

    def getSheet(self, sheetName):
        sheet = self.workbook[sheetName]
        return sheet

    def getRowAndColNum(self, sheet):
        rowNum = sheet.max_row
        colNum = sheet.max_column
        return rowNum, colNum

    def getRowValue(self, sheet, rowNo):
        return list(sheet.rows)[rowNo - 1]

    def getColValue(self, sheet, colNo):
        return list(sheet.columns)[colNo - 1]

    def getCellValue(self, sheet, cellName = 'A1'):
        return sheet[cellName].value

    def getWriteByRowAndCol(self, sheet, content, rowNo, colNo, style=None):
        sheet.cell(row=rowNo, column=colNo).value = content
        if style:
            sheet.cell(row=rowNo, column=colNo).font = Font(color=self.colDict[style])
        self.workbook.save(self.excelFile)

    def getWriteByCell(self, sheet, content, cellName, style = None):
        sheet[cellName].value = content
        if style:
            sheet[cellName].font = Font(color=self.colDict[style])
        self.workbook.save(self.excelFile)

    def getWriteTimeByRowAndCol(self, sheet, rowNo, colNo, nowTime = None, style = None):
        self.now = time.strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row = rowNo, column = colNo).value = nowTime
        if style:
            sheet.cell(row=rowNo, column=colNo).font = Font(color=self.colDict[style])
        self.workbook.save(self.excelFile)

if __name__ == '__main__':
    from actions.PageActions import getOpenLocalFile
    import time

    ep = ExcelParsing()
    excelFile = r'D:\PycharmProjects\KeyWordPlusDataDrivenFramework\testData\携程查询机票酒店.xlsx'
    ep.getWorkBook(excelFile)
    sheet = ep.getSheet('机票查询')
    print(sheet, sheet.title)
    print(ep.getRowAndColNum(sheet))
    # rowValue = ep.getRowValue(sheet, 2)
    # for value in rowValue:
    #     print(value.value, end=', ')
    # for cv in ep.getColValue(sheet, 2):
    #     print(cv.value)
    # print(ep.getCellValue(sheet, 'B2'))
    # ep.getWriteByRowAndCol(sheet, 'test', 2, 1, 'blue')
    # ep.getWriteByCell(sheet, '单元格写入', 'A3', 'red')
    # now = time.strftime('%Y-%m-%d %H:%M:%S')
    ep.getWriteTimeByRowAndCol(sheet,2, 1, ep.now, 'red')
    time.sleep(2)
    ep.getWriteTimeByRowAndCol(sheet, 3, 1, ep.now,'blue')
    getOpenLocalFile(excelFile)



